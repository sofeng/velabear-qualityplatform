from collections import defaultdict
from pathlib import Path
import re

from django.contrib.auth.models import Group
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from apps.users.models import Role, RoleMembership, User


EXPECTED_HEADERS = ('姓', '用户名', '邮箱', '部门', '组别', '职位', '角色', '标签')
TAG_SPLIT_PATTERN = re.compile(r'[\n\r,，、;/；|]+')


def normalize_text(value):
    return str(value or '').strip()


def normalize_tags(value):
    text = normalize_text(value)
    if not text:
        return []

    tags = []
    seen = set()
    for item in TAG_SPLIT_PATTERN.split(text):
        tag = normalize_text(item)
        if not tag or tag in seen:
            continue
        seen.add(tag)
        tags.append(tag)
    return tags


class Command(BaseCommand):
    help = '将成员 Excel 导入到成员、组别、角色列表'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', help='成员信息 Excel 文件路径')
        parser.add_argument(
            '--sheet-name',
            help='工作表名称，默认读取第一个工作表',
        )
        parser.add_argument(
            '--default-password',
            help='新建用户的默认密码；未提供时会为新用户设置不可登录密码',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        excel_path = Path(options['excel_path']).expanduser()
        if not excel_path.exists():
            raise CommandError(f'未找到 Excel 文件: {excel_path}')

        sheet_name = normalize_text(options.get('sheet_name'))
        default_password = options.get('default_password')

        rows = self._load_rows(excel_path, sheet_name=sheet_name or None)
        if not rows:
            raise CommandError('Excel 中没有可导入的数据行')

        imported = self._import_rows(rows, default_password=default_password)
        self._print_summary(imported, excel_path)

    def _load_rows(self, excel_path, sheet_name=None):
        workbook = load_workbook(excel_path, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]

        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, max_col=len(EXPECTED_HEADERS), values_only=True),
            None,
        )
        headers = tuple(normalize_text(item) for item in (header_row or ()))
        if headers != EXPECTED_HEADERS:
            raise CommandError(f'Excel 表头不匹配，期望 {EXPECTED_HEADERS}，实际 {headers}')

        rows = []
        for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            normalized = {
                'full_name': normalize_text(row[0]),
                'username': normalize_text(row[1]),
                'email': normalize_text(row[2]),
                'department': normalize_text(row[3]),
                'group_name': normalize_text(row[4]),
                'position': normalize_text(row[5]),
                'role_name': normalize_text(row[6]),
                'tags': normalize_tags(row[7]),
                'row_number': index,
            }

            if not any(normalized[key] for key in ('full_name', 'username', 'email', 'group_name', 'role_name')):
                continue

            if not normalized['username']:
                raise CommandError(f'第 {index} 行缺少用户名')
            if not normalized['group_name']:
                raise CommandError(f'第 {index} 行缺少组别')
            if not normalized['role_name']:
                raise CommandError(f'第 {index} 行缺少角色')

            rows.append(normalized)
        return rows

    def _import_rows(self, rows, default_password=None):
        summary = {
            'row_count': len(rows),
            'unique_user_count': 0,
            'created_users': 0,
            'updated_users': 0,
            'unchanged_users': 0,
            'created_groups': 0,
            'synced_groups': 0,
            'created_roles': 0,
            'synced_roles': 0,
        }

        group_members = defaultdict(dict)
        role_members = defaultdict(dict)
        user_states = {}

        for row in rows:
            user, created, updated = self._upsert_user(row, default_password=default_password)
            previous_state = user_states.get(user.username)
            next_state = 'created' if created else 'updated' if updated else 'unchanged'
            user_states[user.username] = self._merge_user_state(previous_state, next_state)

            group_members[row['group_name']][user.id] = user
            role_payload = role_members[row['role_name']].setdefault(
                user.id,
                {
                    'user': user,
                    'tags': [],
                },
            )
            role_payload['tags'] = self._merge_tags(role_payload['tags'], row['tags'])

        summary['unique_user_count'] = len(user_states)
        summary['created_users'] = sum(1 for state in user_states.values() if state == 'created')
        summary['updated_users'] = sum(1 for state in user_states.values() if state == 'updated')
        summary['unchanged_users'] = sum(1 for state in user_states.values() if state == 'unchanged')

        for group_name, member_map in group_members.items():
            group, created = Group.objects.get_or_create(name=group_name)
            if created:
                summary['created_groups'] += 1
            desired_members = list(member_map.values())
            group.user_set.set(desired_members)
            summary['synced_groups'] += 1

        for role_name, member_map in role_members.items():
            role, created = Role.objects.get_or_create(name=role_name)
            if created:
                summary['created_roles'] += 1

            desired_user_ids = list(member_map.keys())
            desired_users = [item['user'] for item in member_map.values()]
            role.members.set(desired_users)

            existing_memberships = {
                membership.user_id: membership
                for membership in role.role_memberships.select_related('user').all()
            }

            for user_id, payload in member_map.items():
                membership = existing_memberships.get(user_id)
                tags = payload['tags']
                if membership is None:
                    RoleMembership.objects.create(role=role, user=payload['user'], tags=tags)
                    continue

                if membership.tags != tags:
                    membership.tags = tags
                    membership.save(update_fields=['tags', 'updated_at'])

            role.role_memberships.exclude(user_id__in=desired_user_ids).delete()
            summary['synced_roles'] += 1

        return summary

    def _merge_tags(self, current_tags, next_tags):
        merged_tags = []
        seen = set()

        for tag in list(current_tags or []) + list(next_tags or []):
            normalized_tag = normalize_text(tag)
            if not normalized_tag or normalized_tag in seen:
                continue
            seen.add(normalized_tag)
            merged_tags.append(normalized_tag)

        return merged_tags

    def _merge_user_state(self, current_state, next_state):
        priority = {
            None: 0,
            'unchanged': 1,
            'updated': 2,
            'created': 3,
        }
        return next_state if priority[next_state] >= priority[current_state] else current_state

    def _upsert_user(self, row, default_password=None):
        username = row['username']
        email = row['email']

        user = User.objects.filter(username=username).first()
        email_user = User.objects.filter(email=email).first() if email else None

        if user and email_user and user.id != email_user.id:
            raise CommandError(
                f"用户名 {username} 与邮箱 {email} 分别匹配到不同用户，无法自动导入（Excel 第 {row['row_number']} 行）"
            )

        if user is None and email_user is not None:
            user = email_user

        created = False
        if user is None:
            user = User(username=username)
            created = True

        changed_fields = []
        field_values = {
            'username': username,
            'email': email,
            'first_name': row['full_name'],
            'last_name': '',
            'department': row['department'],
            'position': row['position'],
            'is_active': True,
        }

        for field_name, expected_value in field_values.items():
            current_value = getattr(user, field_name)
            if current_value != expected_value:
                setattr(user, field_name, expected_value)
                changed_fields.append(field_name)

        if created:
            if default_password:
                user.set_password(default_password)
            else:
                user.set_unusable_password()
            user.save()
            return user, True, True

        if changed_fields:
            user.save(update_fields=changed_fields + ['updated_at'])
            return user, False, True

        return user, False, False

    def _print_summary(self, summary, excel_path):
        self.stdout.write(self.style.SUCCESS(f'导入完成: {excel_path}'))
        self.stdout.write(
            '数据行 {row_count}；唯一成员 {unique_user_count}；成员 新建 {created_users} / 更新 {updated_users} / 未变更 {unchanged_users}；'
            '组别 新建 {created_groups} / 同步 {synced_groups}；'
            '角色 新建 {created_roles} / 同步 {synced_roles}'.format(**summary)
        )
