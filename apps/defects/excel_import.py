import re
import uuid
from collections import defaultdict
from html import escape

from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.users.models import User

from .models import Defect

RICH_TEXT_IMAGE_PREFIX = 'defect_rich_text_images/'
TEST_ENVIRONMENT_NAME = '思源测试环境'
HEADER_ALIASES = {
    '责任小组': 'responsibility_group',
    '需求编号': 'requirement_id',
    '模块': 'module',
    '页面': 'page',
    '提交人': 'submitter',
    '问题描述': 'problem_description',
    '预期': 'expected',
    '截图1': 'screenshot_1',
    '截图2': 'screenshot_2',
    '截图3': 'screenshot_3',
    '截图4': 'screenshot_4',
    '优先级': 'priority',
    '解决状态': 'resolution_status',
    '后端': 'backend_developer',
    '前端': 'frontend_developer',
    '问题原因': 'problem_reason',
    '问题根因': 'root_cause',
}
REQUIRED_HEADERS = (
    '责任小组',
    '需求编号',
    '模块',
    '页面',
    '提交人',
    '问题描述',
    '预期',
    '优先级',
    '解决状态',
)
IMAGE_FIELD_KEYS = (
    'expected',
    'screenshot_1',
    'screenshot_2',
    'screenshot_3',
    'screenshot_4',
)
STATUS_RULES = (
    ('打回待处理', 'returned_pending'),
    ('回归验证完成', 'regression_verified'),
    ('待客户环境验证', 'customer_validation'),
    ('待转新需求', 'pending_requirement'),
    ('已转新需求', 'requirement_created'),
    ('暂不处理', 'deferred'),
    ('已作废', 'invalid'),
    ('已关闭', 'closed'),
    ('关闭', 'closed'),
    ('打回', 'returned_pending'),
    ('重新打开', 'reopened'),
    ('重开', 'reopened'),
    ('拒绝', 'rejected'),
    ('驳回', 'rejected'),
    ('作废', 'invalid'),
    ('提测', 'resolved'),
    ('已解决', 'resolved'),
    ('解决', 'resolved'),
    ('处理中', 'in_progress'),
    ('处理', 'in_progress'),
    ('待处理', 'new'),
    ('新建', 'new'),
)


def _normalize_text(value):
    if value is None:
        return ''
    return str(value).replace('\r\n', '\n').replace('\r', '\n').strip()


def _normalize_person_text(value):
    normalized = _normalize_text(value)
    if not normalized:
        return ''

    segments = [
        segment.strip().lstrip('@').strip()
        for segment in re.split(r'[、,，/]+', normalized)
        if segment and segment.strip()
    ]
    if not segments:
        return ''
    return '、'.join(segments)


def _normalize_person_lookup_key(value):
    return re.sub(r'\s+', '', _normalize_person_text(value)).casefold()


def _build_user_name_index():
    name_index = defaultdict(list)

    for user in User.objects.all().order_by('id'):
        candidate_names = {
            _normalize_person_lookup_key(user.username),
            _normalize_person_lookup_key(user.email.split('@', 1)[0] if user.email else ''),
            _normalize_person_lookup_key(user.full_name),
            _normalize_person_lookup_key(f'{user.first_name}{user.last_name}'),
            _normalize_person_lookup_key(f'{user.last_name}{user.first_name}'),
        }
        for candidate_name in candidate_names:
            if candidate_name:
                name_index[candidate_name].append(user)

    return name_index


def _resolve_user_by_name(name, user_name_index):
    lookup_key = _normalize_person_lookup_key(name)
    if not lookup_key:
        return None

    matched_users = user_name_index.get(lookup_key) or []
    return matched_users[0] if matched_users else None


def _normalize_priority(value):
    normalized = _normalize_text(value).upper()
    if normalized in {'P1', 'P2', 'P3', 'P4'}:
        return normalized, None
    if not normalized:
        return 'P3', '缺少优先级，已按 P3 导入'
    return 'P3', f'无法识别的优先级“{normalized}”，已按 P3 导入'


def _normalize_status(value):
    normalized = _normalize_text(value)
    if not normalized:
        return 'new', '缺少解决状态，已按新建导入'

    for keyword, mapped_status in STATUS_RULES:
        if keyword in normalized:
            return mapped_status, None

    return 'new', f'无法识别的解决状态“{normalized}”，已按新建导入'


def _normalize_module_path(value):
    normalized = _normalize_text(value)
    if not normalized:
        return ''

    normalized = normalized.replace('／', '/').replace('\\', '/')
    if '/' in normalized:
        segments = [segment.strip() for segment in re.split(r'\s*/\s*', normalized) if segment.strip()]
    else:
        segments = [segment.strip() for segment in re.split(r'\s*-+\s*', normalized) if segment.strip()]

    return ' / '.join(segments) if segments else normalized


def _build_module_relation_items(module_name, responsibility_group=''):
    normalized_path = _normalize_module_path(module_name)
    if not normalized_path:
        return []

    segments = [segment.strip() for segment in normalized_path.split(' / ') if segment.strip()]
    return [{
        'node_text': segments[-1] if segments else normalized_path,
        'node_type': 'module',
        'path': normalized_path,
        'parent_text': segments[-2] if len(segments) > 1 else '',
        'responsibility_group': _normalize_text(responsibility_group),
    }]


def _build_header_map(worksheet):
    header_map = {}
    for column_index in range(1, worksheet.max_column + 1):
        header_value = _normalize_text(worksheet.cell(row=1, column=column_index).value)
        if header_value:
            header_map[header_value] = column_index
    return header_map


def _build_image_map(worksheet):
    image_map = defaultdict(list)
    for image in getattr(worksheet, '_images', []):
        anchor = getattr(image, 'anchor', None)
        anchor_from = getattr(anchor, '_from', None)
        if anchor_from is None:
            continue
        image_map[(anchor_from.row + 1, anchor_from.col + 1)].append(image)
    return image_map


def _save_rich_text_image(image):
    image_bytes = image._data()
    image_format = str(getattr(image, 'format', '') or 'png').lower()
    if image_format == 'jpg':
        image_format = 'jpeg'
    if image_format not in {'png', 'jpeg', 'gif', 'webp', 'bmp'}:
        image_format = 'png'

    upload_directory = timezone.now().strftime(f'{RICH_TEXT_IMAGE_PREFIX}%Y/%m')
    relative_path = default_storage.save(
        f'{upload_directory}/{uuid.uuid4().hex}.{image_format}',
        ContentFile(image_bytes),
    )
    media_url = f"{settings.MEDIA_URL.rstrip('/')}/{relative_path.lstrip('/')}"
    return media_url, relative_path


def _append_text_paragraphs(chunks, text):
    normalized = _normalize_text(text)
    if not normalized:
        chunks.append('<p><br></p>')
        return

    for line in normalized.split('\n'):
        chunks.append(f'<p>{escape(line) if line else "<br>"}</p>')


def _append_section(chunks, title, *, text='', image_urls=None):
    chunks.append(f'<p>【{escape(title)}】</p>')
    if text:
        _append_text_paragraphs(chunks, text)
    elif not image_urls:
        chunks.append('<p><br></p>')

    for image_url in image_urls or []:
        chunks.append(f'<p><img src="{escape(image_url, quote=True)}"></p>')


def _build_description_html(
    *,
    version_name,
    page,
    problem_description,
    expected_text,
    expected_image_urls,
    actual_text,
    actual_image_urls,
):
    precondition_parts = []
    normalized_page = _normalize_text(page)
    normalized_expected_text = _normalize_text(expected_text)
    if normalized_page:
        precondition_parts.append(normalized_page)
    if normalized_expected_text:
        precondition_parts.append(normalized_expected_text)
    elif expected_image_urls:
        precondition_parts.append('预期结果见下方图片')

    chunks = []
    _append_section(chunks, '版本号', text=_normalize_text(version_name))
    _append_section(chunks, '测试环境', text=TEST_ENVIRONMENT_NAME)
    _append_section(chunks, '测试数据')
    _append_section(chunks, '前置条件', text='\n'.join(precondition_parts))
    _append_section(chunks, '测试步骤', text=problem_description)
    _append_section(chunks, '预期结果', text=expected_text, image_urls=expected_image_urls)
    _append_section(chunks, '实际结果', text=actual_text, image_urls=actual_image_urls)
    return ''.join(chunks)


def _row_has_importable_content(row_data, image_headers):
    if any(row_data.values()):
        return True
    return any(image_headers.values())


def import_defects_from_excel_file(
    *,
    uploaded_file,
    project,
    version,
    operator,
    update_status_metadata,
    create_history,
    record_type=Defect.RECORD_TYPE_DEFECT,
    history_create_field='defect',
):
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header_map = _build_header_map(worksheet)
    missing_headers = [header_name for header_name in REQUIRED_HEADERS if header_name not in header_map]
    if missing_headers:
        raise ValueError(f'Excel 缺少必要列：{", ".join(missing_headers)}')

    image_map = _build_image_map(worksheet)
    user_name_index = _build_user_name_index()

    created_defects = []
    warnings = []
    skipped_count = 0

    for row_number in range(2, worksheet.max_row + 1):
        row_data = {}
        for header_name, field_name in HEADER_ALIASES.items():
            column_index = header_map.get(header_name)
            row_data[field_name] = _normalize_text(
                worksheet.cell(row=row_number, column=column_index).value
            ) if column_index else ''

        image_headers = {
            field_name: image_map.get((row_number, header_map[header_name]), [])
            for header_name, field_name in HEADER_ALIASES.items()
            if field_name in IMAGE_FIELD_KEYS and header_name in header_map
        }

        if not _row_has_importable_content(row_data, image_headers):
            continue

        title = row_data['problem_description']
        if not title:
            skipped_count += 1
            warnings.append(f'第 {row_number} 行缺少问题描述，已跳过')
            continue

        priority, priority_warning = _normalize_priority(row_data['priority'])
        if priority_warning:
            warnings.append(f'第 {row_number} 行：{priority_warning}')

        status, status_warning = _normalize_status(row_data['resolution_status'])
        if status_warning:
            warnings.append(f'第 {row_number} 行：{status_warning}')

        creator_name = row_data['submitter']
        creator_user = _resolve_user_by_name(creator_name, user_name_index)
        if creator_name and creator_user is None:
            warnings.append(f'第 {row_number} 行提交人“{creator_name}”未匹配到平台用户，已使用当前导入人')
        elif not creator_name:
            warnings.append(f'第 {row_number} 行缺少提交人，已使用当前导入人')
        creator_user = creator_user or operator

        saved_paths = []
        try:
            expected_image_urls = []
            for image in image_headers.get('expected', []):
                image_url, relative_path = _save_rich_text_image(image)
                expected_image_urls.append(image_url)
                saved_paths.append(relative_path)

            actual_image_urls = []
            for field_name in ('screenshot_1', 'screenshot_2', 'screenshot_3', 'screenshot_4'):
                for image in image_headers.get(field_name, []):
                    image_url, relative_path = _save_rich_text_image(image)
                    actual_image_urls.append(image_url)
                    saved_paths.append(relative_path)

            actual_text = '\n'.join(
                filter(
                    None,
                    [
                        row_data['screenshot_1'],
                        row_data['screenshot_2'],
                        row_data['screenshot_3'],
                        row_data['screenshot_4'],
                    ],
                )
            )
            description = _build_description_html(
                version_name=version.name,
                page=row_data['page'],
                problem_description=title,
                expected_text=row_data['expected'],
                expected_image_urls=expected_image_urls,
                actual_text=actual_text,
                actual_image_urls=actual_image_urls,
            )

            frontend_developer = _normalize_person_text(row_data['frontend_developer'])
            backend_developer = _normalize_person_text(row_data['backend_developer'])
            module_items = _build_module_relation_items(
                row_data['module'],
                responsibility_group=row_data['responsibility_group'],
            )

            with transaction.atomic():
                defect = Defect.objects.create(
                    record_type=record_type,
                    project=project,
                    version=version,
                    title=title,
                    description=description,
                    problem_reason=row_data['problem_reason'],
                    root_cause=row_data['root_cause'],
                    priority=priority,
                    severity='medium',
                    status=status,
                    requirement_id=row_data['requirement_id'],
                    modules=module_items,
                    related_testcases=[],
                    related_testpoints=[],
                    labels=[],
                    created_by=creator_user,
                    frontend_developer=frontend_developer,
                    backend_developer=backend_developer,
                )

                update_status_metadata(defect, creator_user, defect.status)
                defect.save()

                create_history(
                    defect,
                    changed_by=operator,
                    field=history_create_field,
                    action='create',
                    from_value=None,
                    to_value={
                        'code': defect.code,
                        'title': defect.title,
                        'status': defect.status,
                    },
                )

            created_defects.append(defect)
        except Exception:
            for relative_path in saved_paths:
                if default_storage.exists(relative_path):
                    default_storage.delete(relative_path)
            raise

    return {
        'created_count': len(created_defects),
        'skipped_count': skipped_count,
        'warning_count': len(warnings),
        'warnings': warnings,
        'created_records': [
            {
                'id': defect.id,
                'code': defect.code,
                'title': defect.title,
            }
            for defect in created_defects
        ],
    }
