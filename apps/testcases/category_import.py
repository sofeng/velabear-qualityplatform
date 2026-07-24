from collections import OrderedDict
from pathlib import Path

from django.conf import settings
from django.db import transaction
from openpyxl import load_workbook

from .models import ManualTestCaseCategory

DEFAULT_ROOT_NAME = '物业通'
HEADER_LEVEL_1 = '一级菜单'
HEADER_LEVEL_2 = '二级菜单'


def resolve_excel_path(excel_path):
    path = Path(excel_path).expanduser()
    if path.exists():
        return path

    project_relative_path = Path(settings.BASE_DIR) / excel_path
    if project_relative_path.exists():
        return project_relative_path

    raise FileNotFoundError(f'Excel文件不存在: {excel_path}')


def _normalize_cell(value):
    return str(value or '').strip()


def _find_header_row(sheet):
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        first = _normalize_cell(row[0] if len(row) > 0 else '')
        second = _normalize_cell(row[1] if len(row) > 1 else '')
        if first == HEADER_LEVEL_1 and second == HEADER_LEVEL_2:
            return index
    return 1


def extract_manual_category_tree(excel_path, sheet_name=None):
    resolved_path = resolve_excel_path(excel_path)
    workbook = load_workbook(resolved_path, read_only=True, data_only=True)

    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        header_row = _find_header_row(worksheet)
        tree = OrderedDict()

        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            level_1 = _normalize_cell(row[0] if len(row) > 0 else '')
            level_2 = _normalize_cell(row[1] if len(row) > 1 else '')

            if not level_1:
                continue

            children = tree.setdefault(level_1, [])
            if level_2 and level_2 not in children:
                children.append(level_2)

        return tree
    finally:
        workbook.close()


@transaction.atomic
def import_manual_categories_from_excel(excel_path, project, root_name=DEFAULT_ROOT_NAME, sheet_name=None):
    category_tree = extract_manual_category_tree(excel_path, sheet_name=sheet_name)
    if not category_tree:
        raise ValueError('Excel中未解析到可导入的一级菜单数据')

    root_category, root_created = ManualTestCaseCategory.objects.get_or_create(
        project=project,
        parent=None,
        name=root_name,
        defaults={
            'description': '手工用例根目录',
            'order': 0,
        },
    )

    created_level_1 = 0
    created_level_2 = 0
    updated_level_1 = 0
    updated_level_2 = 0

    for level_1_order, (level_1_name, level_2_children) in enumerate(category_tree.items(), start=1):
        category, created = ManualTestCaseCategory.objects.get_or_create(
            project=project,
            parent=root_category,
            name=level_1_name,
            defaults={
                'description': '',
                'order': level_1_order,
            },
        )
        if created:
            created_level_1 += 1
        elif category.order != level_1_order:
            category.order = level_1_order
            category.save(update_fields=['order', 'updated_at'])
            updated_level_1 += 1

        for level_2_order, level_2_name in enumerate(level_2_children, start=1):
            child, child_created = ManualTestCaseCategory.objects.get_or_create(
                project=project,
                parent=category,
                name=level_2_name,
                defaults={
                    'description': '',
                    'order': level_2_order,
                },
            )
            if child_created:
                created_level_2 += 1
            elif child.order != level_2_order:
                child.order = level_2_order
                child.save(update_fields=['order', 'updated_at'])
                updated_level_2 += 1

    return {
        'root_category_id': root_category.id,
        'root_created': root_created,
        'level_1_total': len(category_tree),
        'level_2_total': sum(len(children) for children in category_tree.values()),
        'created_level_1': created_level_1,
        'created_level_2': created_level_2,
        'updated_level_1': updated_level_1,
        'updated_level_2': updated_level_2,
    }
